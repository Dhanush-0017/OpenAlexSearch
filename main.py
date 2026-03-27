import os
import sys
import json
import glob
import textwrap
import openpyxl
from openpyxl.styles import Font, Alignment

from pipeline.chunker    import process_pdf
from pipeline.summarizer import process_paper as summarize_paper

PAPERS_DIR  = "papers"
RESULTS_DIR = "results"
HEADERS     = ["title", "summary", "classification", "accuracy", "justification"]


def get_pdf_files():
    pdfs = glob.glob(os.path.join(PAPERS_DIR, "*.pdf"))
    if not pdfs:
        print(f"No PDF files found in '{PAPERS_DIR}/' folder.")
        sys.exit(1)
    return pdfs


def clear_results():
    """Delete old results files at the start of each run."""
    os.makedirs(RESULTS_DIR, exist_ok=True)
    for f in ["results.xlsx", "results.json"]:
        path = os.path.join(RESULTS_DIR, f)
        if os.path.isfile(path):
            os.remove(path)


def justify_text(text, line_width, initial_indent, subsequent_indent):
    """Full (both-side) justification for terminal output."""
    wrapped = textwrap.wrap(text, width=line_width - len(subsequent_indent))
    lines   = []
    for i, line in enumerate(wrapped):
        indent = initial_indent if i == 0 else subsequent_indent
        # Last line: left-align only
        if i == len(wrapped) - 1:
            lines.append(indent + line)
            continue
        words = line.split()
        if len(words) == 1:
            lines.append(indent + line)
            continue
        total_fill  = line_width - len(indent) - sum(len(w) for w in words)
        gaps        = len(words) - 1
        base, extra = divmod(total_fill, gaps)
        justified   = ""
        for j, word in enumerate(words[:-1]):
            justified += word + " " * (base + (1 if j < extra else 0))
        justified  += words[-1]
        lines.append(indent + justified)
    return "\n".join(lines)


def save_results(data):
    os.makedirs(RESULTS_DIR, exist_ok=True)

    xlsx_path = os.path.join(RESULTS_DIR, "results.xlsx")
    json_path = os.path.join(RESULTS_DIR, "results.json")

    classification = data.get("classification", {})

    summary       = "\n".join(textwrap.wrap(data.get("summary", ""),       width=100))
    justification = "\n".join(textwrap.wrap(classification.get("justification", ""), width=100))

    row = [
        data["title"],
        summary,
        classification.get("category", ""),
        classification.get("confidence", "") + "%",
        justification,
    ]

    # ── Excel ─────────────────────────────────────────────────────────────────
    wrap = Alignment(wrap_text=True, vertical="top")
    try:
        if os.path.isfile(xlsx_path):
            wb = openpyxl.load_workbook(xlsx_path)
            ws = wb.active
        else:
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.append(HEADERS)
            bold = Font(bold=True)
            for cell in ws[1]:
                cell.font      = bold
                cell.alignment = Alignment(vertical="center")
        ws.append(row)
        # Apply wrap text to summary (col 2) and justification (col 5)
        last_row = ws.max_row
        ws.cell(last_row, 2).alignment = wrap
        ws.cell(last_row, 5).alignment = wrap
        wb.save(xlsx_path)
    except PermissionError:
        print(f"  Could not write Excel — close {xlsx_path} and try again.")

    # ── JSON ──────────────────────────────────────────────────────────────────
    existing = []
    if os.path.isfile(json_path):
        try:
            with open(json_path, "r", encoding="utf-8") as f:
                existing = json.load(f)
        except (json.JSONDecodeError, IOError):
            existing = []

    existing.append({
        "title"         : data["title"],
        "num_chunks"    : data["num_chunks"],
        "final_summary" : data.get("summary", ""),
        "classification": classification,
    })

    with open(json_path, "w", encoding="utf-8") as f:
        json.dump(existing, f, indent=2, ensure_ascii=False)


def process(pdf_path):
    paper_id = os.path.splitext(os.path.basename(pdf_path))[0]

    print(f"\n{'='*80}")
    print(f"  {os.path.basename(pdf_path)}")
    print(f"{'='*80}")

    print("  Chunking...", end="", flush=True)
    chunks, title = process_pdf(pdf_path, paper_id=paper_id)
    print(f" {len(chunks)} chunks extracted")

    print("  Summarizing & Classifying...", end="", flush=True)
    data = summarize_paper(paper_id)
    print(" done")

    classification = data.get("classification", {})
    save_results(data)

    justification = justify_text(
        classification['justification'],
        line_width        = 80,
        initial_indent    = "  Justification           : ",
        subsequent_indent = "                            ",
    )
    print(f"\n  Title                   : {title}")
    print(f"  Category                : {classification['category']}")
    print(f"  Classification Accuracy : {classification['confidence']}%")
    print(justification)


def main():
    print("=" * 50)
    print("  Research Paper Analysis   ")
    print("=" * 50)

    pdfs = get_pdf_files()
    clear_results()
    print(f"\n  Found {len(pdfs)} Research Paper(s) in '{PAPERS_DIR}/'")

    for pdf_path in pdfs:
        process(pdf_path)

    print(f"\n{'='*80}")
    print("  All papers processed.")
    print(f"  Saved → results/results.xlsx & results/results.json")
    print(f"{'='*80}")


if __name__ == "__main__":
    main()
